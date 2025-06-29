#/bin/python3
import pdbufr
import pandas as pd
import numpy as np 
from openpyxl import load_workbook
from openpyxl.styles import Border, Side


bufr_18 = '/home/lagha/data/bufr/synop_alg_202402281800.bufr'
bufr_06 = '/home/lagha/data/bufr/synop_alg_202402290600.bufr'


df_06 = pdbufr.read_bufr(bufr_06, columns=("stationOrSiteName", 
                                          #"stationNumber",
                                          "heightOfStationGroundAboveMeanSeaLevel",
                                          "windDirection",
                                          "windSpeed",
                                          #"airTemperature",
                                          "cloudAmount",
                                          "totalPrecipitationOrTotalWaterEquivalent",
                                          "maximumTemperatureAtHeightAndOverPeriodSpecified",
                                          "minimumTemperatureAtHeightAndOverPeriodSpecified"
                                          ))
df_18 = pdbufr.read_bufr(bufr_18, columns=("stationOrSiteName",                                        
                                        "maximumTemperatureAtHeightAndOverPeriodSpecified"
                                        ))

df_06.rename(columns={"stationOrSiteName":"STATIONS",                                     
                   "heightOfStationGroundAboveMeanSeaLevel":"ALTITUDE EN METRES",
                   "windDirection":"Dir","windSpeed":"Vit (m/s)",
                   "cloudAmount":"Néb (1/8)",
                   "totalPrecipitationOrTotalWaterEquivalent":"Précip (mm)",
                   "maximumTemperatureAtHeightAndOverPeriodSpecified":"Max de la veille",
                   "minimumTemperatureAtHeightAndOverPeriodSpecified":"Min de la nuit"
                   }, inplace=True)

          
df_06["Max de la veille"] = (df_18["maximumTemperatureAtHeightAndOverPeriodSpecified"] - 273.15).round(1)
df_06["Min de la nuit"] = (df_06["Min de la nuit"] - 273.15).round(1)


#df_06["Dir"] = df_06["Dir"]
#df_06["Vit (m/s)"] = df_06["Vit (m/s)"]
#df_06["Vit (m/s)"] = pd.to_numeric(df_06["Vit (m/s)"], errors='coerce')
#df_06["Vit (m/s)"].fillna(0, inplace=True)

df_06 = df_06.dropna(subset = ["ALTITUDE EN METRES"])
df_06["Phénoménes"] = None
df_06 = df_06.sort_values(["STATIONS"])
df_06 = df_06[["STATIONS","ALTITUDE EN METRES","Dir","Vit (m/s)","Néb (1/8)","Phénoménes","Précip (mm)","Phénoménes","Max de la veille","Min de la nuit"]]


# Define rose wind directions and sectors for 8 directions
rose_wind = {
    'N': list(range(338, 361)) + list(range(0, 23)),
    'NE': list(range(23, 68)),
    'E': list(range(68, 113)),
    'SE': list(range(113, 158)),
    'S': list(range(158, 203)),
    'SW': list(range(203, 248)),
    'W': list(range(248, 293)),
    'NW': list(range(293, 338))
}

# Function to convert degrees to 8-point rose wind
def convert_to_wind_rose(deg):
    for direction, sector in rose_wind.items():
        if int(deg) in sector:
            return direction

# Apply conversion and conditions to a DataFrame column 'Dir'
df_06["Dir"] = df_06["Dir"].apply(convert_to_wind_rose)

df_06.loc[df_06["Vit (m/s)"] == 0, "Dir"] = 'Calme'
df_06.loc[(df_06["Vit (m/s)"] < 2) & (df_06["Vit (m/s)"] > 0), "Dir"] = 'VRB'

#df_06.to_csv('output_06.csv', index=True)

# Load Excel template
template_path = 'template.xlsx'
wb = load_workbook(template_path)
ws = wb.active

# Insert DataFrame into Excel template starting from 4th row
start_row = 4
for row in df_06.itertuples(index=False):
    ws.append(row)

# Define border style
border_style = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

# Apply border to all cells
for row in ws.iter_rows():
    
    for cell in row:
        cell.border = border_style

# Save to Excel
wb.save('output1.xlsx')


import pandas as pd
from openpyxl import load_workbook

# Load data from output1.xlsx and list_stations.xlsx
output_df = pd.read_excel('output1.xlsx')
list_stations_wb = load_workbook('liste_stations.xlsx')
list_stations_ws = list_stations_wb.active

# Iterate over list_stations.xlsx and search for each station in output1.xlsx
for row_idx, row in enumerate(list_stations_ws.iter_rows(), start=1):
    station_name = row[0].value  # Assuming STATIONS is in the first column
    if station_name:  # Check if station name is not empty
        # Find the row index of the station in output1.xlsx
        station_row_index = output_df.index[output_df['STATIONS'] == station_name].tolist()
        if station_row_index:
            # Write the data of the station from output1.xlsx to list_stations.xlsx
            station_row_index = station_row_index[0]  # Assuming there's only one match
            station_data = output_df.iloc[station_row_index].tolist()
            for col_idx, cell in enumerate(row, start=1):
                cell.value = station_data[col_idx - 1]

# Save the modified list_stations.xlsx
list_stations_wb.save('liste_stations1.xlsx')

